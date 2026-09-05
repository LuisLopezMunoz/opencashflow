"""Export a computed cashflow matrix to CSV or XLSX.

Split out of cli.py to keep that file from growing unbounded — this
is the one piece with real complexity (two-pass layout, rule-to-formula
translation), everything else in the CLI is a thin wrapper over the engine
or a handful of ORM queries.
"""
import csv
import sys
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple

from opencashflow.models import CellOverride, SheetCell, SheetRow, SheetSection


def _period_label(period) -> str:
    return period.label or period.period_date.strftime("%b %Y")


def _scaled(value, unit: str):
    if value is None:
        return None
    divisor = Decimal(1000) if unit == "k" else Decimal(1)
    return float((Decimal(value) / divisor).to_integral_value(rounding=ROUND_HALF_UP))


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_csv(sheet, result, periods, unit: str, show_ids: bool, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Fila"] + [_period_label(p) for p in periods])
        for section in result["sections"]:
            for row_data in section["rows"]:
                row = row_data["row"]
                label = f"[{row.id}] {row.name}" if show_ids else row.name
                cells_by_period = {c.period_id: c for c in row_data["cells"]}
                values = []
                for p in periods:
                    cell = cells_by_period.get(p.id)
                    scaled = _scaled(cell.projected_value, unit) if cell else None
                    values.append("" if scaled is None else str(scaled))
                writer.writerow([label] + values)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def _rule_to_formula(
    rule: Optional[Dict],
    row_id: int,
    period_id: int,
    row_id_to_excel_row: Dict[int, int],
    period_id_to_col_letter: Dict[int, str],
    sorted_period_ids: List[int],
    row_signs: Dict[int, int],
) -> Optional[str]:
    """Translate a projection rule into an Excel formula string (starting
    with '='), or None if it can't be represented with what's in the
    exported range (caller falls back to a literal + explanatory comment).
    `constant` is deliberately never turned into a formula -- a literal
    number is simpler and exactly as correct.
    """
    if not rule:
        return None
    rule_type = rule.get("type")
    col = period_id_to_col_letter.get(period_id)

    if rule_type == "sum_rows":
        terms = []
        for rid in rule.get("row_ids", []):
            if rid not in row_id_to_excel_row or col is None:
                return None
            excel_row = row_id_to_excel_row[rid]
            sign = row_signs.get(rid, 1)
            terms.append(f"{'+' if sign > 0 else '-'}{col}{excel_row}")
        if not terms:
            return None
        formula = "".join(terms)
        if formula.startswith("+"):
            formula = formula[1:]
        return "=" + formula

    if rule_type == "previous_period":
        if period_id not in sorted_period_ids:
            return None
        idx = sorted_period_ids.index(period_id)
        if idx == 0:
            return None
        prev_period_id = sorted_period_ids[idx - 1]
        prev_col = period_id_to_col_letter.get(prev_period_id)
        source_row_id = rule.get("row_id", row_id)
        excel_row = row_id_to_excel_row.get(source_row_id)
        if prev_col is None or excel_row is None:
            return None
        return f"={prev_col}{excel_row}"

    if rule_type == "percent_of_row":
        source_row_id = rule.get("row_id")
        percent = rule.get("percent")
        excel_row = row_id_to_excel_row.get(source_row_id)
        if excel_row is None or col is None or percent is None:
            return None
        return f"={col}{excel_row}*{percent}/100"

    if rule_type == "rolling_average":
        n = rule.get("n")
        excel_row = row_id_to_excel_row.get(row_id)
        if not n or excel_row is None or period_id not in sorted_period_ids:
            return None
        idx = sorted_period_ids.index(period_id)
        start_idx = max(0, idx - n)
        if start_idx == idx:
            return None
        start_col = period_id_to_col_letter.get(sorted_period_ids[start_idx])
        end_col = period_id_to_col_letter.get(sorted_period_ids[idx - 1])
        if not start_col or not end_col:
            return None
        return f'=IFERROR(AVERAGE({start_col}{excel_row}:{end_col}{excel_row}),"")'

    return None  # constant / empty / unsupported -> caller uses a literal


def _cell_content(
    cell_result,
    row: SheetRow,
    period_id: int,
    active_override: Optional[CellOverride],
    row_id_to_excel_row: Dict[int, int],
    period_id_to_col_letter: Dict[int, str],
    sorted_period_ids: List[int],
    row_signs: Dict[int, int],
    mode: str,
) -> Tuple[Optional[object], Optional[str]]:
    """Return (value_or_formula, comment). In "values" mode this is always a
    plain number (or None). In "formulas" mode it's a formula string
    (starting with '=') when the rule can be represented within the
    exported range, otherwise a literal fallback with an explanatory
    comment -- never a dangling reference.
    """
    projected = float(cell_result.projected_value) if cell_result and cell_result.projected_value is not None else None

    if mode == "values":
        return (projected, None)

    if active_override is not None:
        if active_override.override_type in ("manual_value", "lock"):
            val = float(active_override.value) if active_override.value is not None else None
            return (val, active_override.note)
        if active_override.override_type == "manual_rule":
            formula = _rule_to_formula(
                active_override.custom_rule, row.id, period_id,
                row_id_to_excel_row, period_id_to_col_letter, sorted_period_ids, row_signs,
            )
            if formula:
                return (formula, active_override.note)
            note = "Regla del override fuera del rango exportado -- valor ya calculado, no fórmula."
            if active_override.note:
                note += f" ({active_override.note})"
            return (projected, note)

    if cell_result and cell_result.error:
        return (None, f"Error del motor: {cell_result.error}")

    rule = row.default_projection_rule
    if not rule or rule.get("type") in (None, "empty"):
        return (None, None)
    if rule.get("type") == "constant":
        return (rule.get("value"), None)

    formula = _rule_to_formula(rule, row.id, period_id, row_id_to_excel_row, period_id_to_col_letter,
                                sorted_period_ids, row_signs)
    if formula:
        return (formula, None)
    return (projected, "Fuera del rango exportado (referencia a un período no incluido) -- valor ya calculado, no fórmula.")


def export_xlsx(
    sheet, result, periods, all_rows_by_id: Dict[int, SheetRow],
    unit: str, mode: str, styled: bool, show_ids: bool, db, path: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Falta la dependencia 'openpyxl'. Instálala (ya está en requirements.txt) para exportar a xlsx.",
              file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet.name[:31] or "Flujo de Caja"  # Excel sheet-name length limit

    row_signs = {r.id: (-1 if r.sign == "negative" else 1) for r in all_rows_by_id.values()}
    sorted_period_ids = [p.id for p in periods]

    # Active overrides for every cell in this sheet, keyed by (row_id, period_id) --
    # compute_sheet()'s CellResult collapses manual_value/manual_rule/lock into
    # a single "manual" source, which isn't enough detail for formula mode.
    override_rows = (
        db.query(CellOverride, SheetCell.row_id, SheetCell.period_id)
        .join(SheetCell, CellOverride.cell_id == SheetCell.id)
        .join(SheetRow, SheetCell.row_id == SheetRow.id)
        .join(SheetSection, SheetRow.section_id == SheetSection.id)
        .filter(SheetSection.sheet_id == sheet.id, CellOverride.superseded_at.is_(None))
        .all()
    )
    active_override_by_cell = {(rid, pid): ov for ov, rid, pid in override_rows}

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E6F5", end_color="D9E6F5", fill_type="solid") if styled else None
    subtotal_font = Font(bold=True) if styled else None

    # --- Pass 1: layout -- header row, row-label column, and the two
    # mappings a formula needs (rules can reference rows that appear later). ---
    ws.cell(row=1, column=1, value="Fila")
    if styled:
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill

    period_id_to_col_letter: Dict[int, str] = {}
    for i, period in enumerate(periods):
        col = i + 2  # column A is the row label
        cell = ws.cell(row=1, column=col, value=_period_label(period))
        if styled:
            cell.font = header_font
            cell.fill = header_fill
        period_id_to_col_letter[period.id] = get_column_letter(col)

    row_id_to_excel_row: Dict[int, int] = {}
    excel_row_num = 2
    ordered_rows: List[SheetRow] = []
    for section in result["sections"]:
        for row_data in section["rows"]:
            row = row_data["row"]
            ordered_rows.append(row)
            label = f"[{row.id}] {row.name}" if show_ids else row.name
            cell = ws.cell(row=excel_row_num, column=1, value=label)
            if styled and row.row_type in ("subtotal", "total", "running_balance"):
                cell.font = subtotal_font
            row_id_to_excel_row[row.id] = excel_row_num
            excel_row_num += 1

    # --- Pass 2: content, now that both mappings are complete. ---
    for section in result["sections"]:
        for row_data in section["rows"]:
            row = row_data["row"]
            excel_row = row_id_to_excel_row[row.id]
            cells_by_period = {c.period_id: c for c in row_data["cells"]}
            for period in periods:
                col_letter = period_id_to_col_letter[period.id]
                col_idx = periods.index(period) + 2
                cell_result = cells_by_period.get(period.id)
                active_ov = active_override_by_cell.get((row.id, period.id))

                value, comment = _cell_content(
                    cell_result, row, period.id, active_ov,
                    row_id_to_excel_row, period_id_to_col_letter, sorted_period_ids, row_signs, mode,
                )

                xcell = ws.cell(row=excel_row, column=col_idx)
                if isinstance(value, str) and value.startswith("="):
                    xcell.value = value
                elif value is not None:
                    scaled = _scaled(value, unit)
                    xcell.value = scaled
                    xcell.number_format = "#,##0"
                if styled and row.row_type in ("subtotal", "total", "running_balance"):
                    xcell.font = subtotal_font
                if comment:
                    xcell.comment = Comment(comment, "opencashflow")

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 32
    for col_letter in period_id_to_col_letter.values():
        ws.column_dimensions[col_letter].width = 14

    wb.save(path)
